# The following code is used to watch a video stream, detect Aruco markers, and use
# a set of markers to determine the posture of the camera in relation to the plane
# of markers.

import numpy as np
import cv2
import cv2.aruco as aruco
import sys
import os
import pickle
import math
import pandas as pd
import xlsxwriter
from openpyxl import Workbook #pip install openpyxl
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import gspread
from oauth2client.service_account import ServiceAccountCredentials



# Check for camera calibration data
if not os.path.exists('./calibration/CameraCalibration.pckl'):
    print("You need to calibrate the camera you'll be using. See calibration project directory for details.")
    exit()
else:
    f = open('./calibration/CameraCalibration.pckl', 'rb')
    (cameraMatrix, distCoeffs, _, _) = pickle.load(f)
    f.close()
    if cameraMatrix is None or distCoeffs is None:
        print("Calibration issue. Remove ./calibration/CameraCalibration.pckl and recalibrate your camera with calibration_ChAruco.py.")
        exit()




def isRotationMatrix(R):
    Rt = np.transpose(R)
    shouldBeIdentity = np.dot(Rt, R)
    I = np.identity(3, dtype=R.dtype)
    n = np.linalg.norm(I - shouldBeIdentity)
    return n < 1e-6


# Calculates rotation matrix to euler angles
# The result is the same as MATLAB except the order
# of the euler angles ( x and z are swapped ).
def rotationMatrixToEulerAngles(R):
    assert (isRotationMatrix(R))

    sy = math.sqrt(R[0, 0] * R[0, 0] + R[1, 0] * R[1, 0])

    singular = sy < 1e-6

    if not singular:
        x = math.atan2(R[2, 1], R[2, 2])
        y = math.atan2(-R[2, 0], sy)
        z = math.atan2(R[1, 0], R[0, 0])
    else:
        x = math.atan2(-R[1, 2], R[1, 1])
        y = math.atan2(-R[2, 0], sy)
        z = 0

    return np.array([x, y, z])

R_flip  = np.zeros((3,3), dtype=np.float32)
R_flip[0,0] = 1.0
R_flip[1,1] =-1.0
R_flip[2,2] =-1.0

# Constant parameters used in Aruco methods
ARUCO_PARAMETERS = aruco.DetectorParameters_create()
ARUCO_DICT = aruco.Dictionary_get(aruco.DICT_4X4_100)

# Create grid board object we're using in our stream
board = aruco.GridBoard_create(
        markersX=1,
        markersY=1,
        markerLength=0.09,
        markerSeparation=0.01,
        dictionary=ARUCO_DICT)

# Create vectors we'll be using for rotations and translations for postures
rotation_vectors, translation_vectors = None, None
axis = np.float32([[-.5,-.5,0], [-.5,.5,0], [.5,.5,0], [.5,-.5,0],
                   [-.5,-.5,1],[-.5,.5,1],[.5,.5,1],[.5,-.5,1] ])

# Make output image fullscreen
cv2.namedWindow('frame',cv2.WINDOW_NORMAL)


cam = cv2.VideoCapture(0)
cam.set(cv2.CAP_PROP_AUTOFOCUS, 0)
cam.set(cv2.CAP_PROP_FRAME_WIDTH, 1280)
cam.set(cv2.CAP_PROP_FRAME_HEIGHT, 720)

font = cv2.FONT_HERSHEY_PLAIN
while(cam.isOpened()):
    # Capturing each frame of our video stream
    ret, frame = cam.read()
    if ret == True:
        # grayscale image
        gray = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)
        
        # Detect Aruco markers  
        corners, ids, rejectedImgPoints = aruco.detectMarkers(gray, ARUCO_DICT, parameters=ARUCO_PARAMETERS)
  
        # Refine detected markers
        # Eliminates markers not part of our board, adds missing markers to the board
        corners, ids, rejectedImgPoints, recoveredIds = aruco.refineDetectedMarkers(
                image = gray,
                board = board,
                detectedCorners = corners,
                detectedIds = ids,
                rejectedCorners = rejectedImgPoints,
                cameraMatrix = cameraMatrix,
                distCoeffs = distCoeffs)


        # Outline all of the markers detected in our image
        # Uncomment below to show ids as well
        # ProjectImage = aruco.drawDetectedMarkers(ProjectImage, corners, ids, borderColor=(0, 0, 255))
        frame = aruco.drawDetectedMarkers(frame, corners, borderColor=(0, 0, 255))

        # Draw the Charuco board we've detected to show our calibrator the board was properly detected
        # Require at least 1 marker before drawing axis
        if ids is not None and len(ids) > 0:
            # Estimate the posture per each Aruco marker

            if ids is not None and ids[0] == 40:
                write_wb = Workbook()
                rotation_vectors, translation_vectors, _objPoints = aruco.estimatePoseSingleMarkers(corners, 1, cameraMatrix, distCoeffs)

                ret = aruco.estimatePoseSingleMarkers(corners,1, cameraMatrix, distCoeffs)
                rvec, tvec = ret[0][0,0,:], ret[1][0,0,:]
                

                
                aruco.drawDetectedMarkers(frame, corners)
                aruco.drawAxis(frame, cameraMatrix, distCoeffs, rvec, tvec, 1)

                str_position = "MARKER Position x=%4.0f  y=%4.0f  z=%4.0f"%(tvec[0], tvec[1], tvec[2])
                cv2.putText(frame, str_position, (0, 100), font, 1, (0, 255, 0), 2, cv2.LINE_AA)

                write_ws = write_wb.create_sheet('id40',0)
                write_ws['A1'] = 'x'
                write_ws['B1'] = 'y'
                write_ws['C1'] = 'z'

                write_ws.append([tvec[0], tvec[1], tvec[2]])
                write_wb.save('c:\ss\s\data\id40.xlsx')
                                    
            
        cv2.imshow('frame', frame)

    # Exit at the end of the video on the 'q' keypress
    if cv2.waitKey(1) & 0xFF == ord('q'):
        break



cv2.destroyAllWindows()
